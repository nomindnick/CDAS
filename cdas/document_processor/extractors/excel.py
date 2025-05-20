"""
Excel document extractor module.
"""
import re
from typing import Dict, List, Any, Union, Optional, Tuple
from pathlib import Path
import logging
from datetime import datetime

import pandas as pd
import numpy as np
from openpyxl import load_workbook

from cdas.document_processor.processor import BaseExtractor

logger = logging.getLogger(__name__)


class ExcelExtractor(BaseExtractor):
    """
    Extractor for Excel documents.
    """
    
    def __init__(self):
        """Initialize the Excel extractor."""
        pass
    
    def extract_data(self, file_path: Union[str, Path], **kwargs) -> Dict[str, Any]:
        """
        Extract data from an Excel document.
        
        Args:
            file_path: Path to the Excel document
            **kwargs: Additional extractor-specific arguments
                - sheet_name: Specific sheet to extract (optional)
            
        Returns:
            Dictionary containing extracted data
        """
        logger.info(f"Extracting data from Excel: {file_path}")
        
        # Convert to Path object if string
        if isinstance(file_path, str):
            file_path = Path(file_path)
        
        # Check if specific sheet requested
        sheet_name = kwargs.get('sheet_name', None)
        
        # Basic extracted data structure
        extracted_data = {
            "sheets": [],
            "summary": {},
        }
        
        try:
            # First use openpyxl to get workbook properties
            wb = load_workbook(file_path, read_only=True, data_only=True)
            
            # Get sheet names
            sheet_names = wb.sheetnames
            
            # Get document properties
            extracted_data["summary"] = {
                "sheet_count": len(sheet_names),
                "sheet_names": sheet_names,
                "document_properties": self._extract_workbook_properties(wb),
            }
            
            # Close workbook to free resources
            wb.close()
            
            # Process sheets with pandas for data extraction
            if sheet_name:
                # Process single sheet if specified
                if sheet_name in sheet_names:
                    sheet_data = self._extract_sheet_data(file_path, sheet_name)
                    extracted_data["sheets"].append(sheet_data)
                else:
                    logger.warning(f"Requested sheet '{sheet_name}' not found in workbook")
            else:
                # Process all sheets
                for sheet in sheet_names:
                    sheet_data = self._extract_sheet_data(file_path, sheet)
                    extracted_data["sheets"].append(sheet_data)
        
        except Exception as e:
            logger.error(f"Error extracting data from Excel {file_path}: {str(e)}")
            raise
        
        return extracted_data
    
    def extract_line_items(self, file_path: Union[str, Path], **kwargs) -> List[Dict[str, Any]]:
        """
        Extract financial line items from an Excel document.
        
        Args:
            file_path: Path to the Excel document
            **kwargs: Additional extractor-specific arguments
                - sheet_name: Specific sheet to extract (optional)
            
        Returns:
            List of dictionaries, each representing a line item
        """
        logger.info(f"Extracting line items from Excel: {file_path}")
        
        # Extract all data first
        extracted_data = self.extract_data(file_path, **kwargs)
        
        # Initialize line items list
        line_items = []
        
        # Process each sheet to find financial tables
        for sheet_data in extracted_data["sheets"]:
            sheet_name = sheet_data["name"]
            tables = sheet_data["tables"]
            
            for table_idx, table in enumerate(tables):
                # Skip small tables or empty tables
                if table["row_count"] <= 1 or table["column_count"] <= 1:
                    continue
                
                # Convert table data to pandas DataFrame for easier processing
                df = pd.DataFrame(table["data"])
                
                # Skip if DataFrame is empty
                if df.empty:
                    continue
                
                # Detect if table has headers and potentially replace first row with headers
                if table["has_header"] and len(df) > 0:
                    df.columns = df.iloc[0]
                    df = df.iloc[1:].reset_index(drop=True)
                
                # Try to identify if this is a financial table
                if self._is_financial_dataframe(df):
                    # Extract financial line items
                    table_line_items = self._extract_financial_items_from_dataframe(
                        df, sheet_name, table_idx
                    )
                    line_items.extend(table_line_items)
        
        return line_items
    
    def extract_metadata(self, file_path: Union[str, Path], **kwargs) -> Dict[str, Any]:
        """
        Extract metadata from an Excel document.
        
        Args:
            file_path: Path to the Excel document
            **kwargs: Additional extractor-specific arguments
            
        Returns:
            Dictionary containing document metadata
        """
        logger.info(f"Extracting metadata from Excel: {file_path}")
        
        # Convert to Path object if string
        if isinstance(file_path, str):
            file_path = Path(file_path)
        
        # Initialize metadata dictionary
        metadata = {
            "filename": file_path.name,
            "file_size": file_path.stat().st_size,
            "creation_date": None,
            "modification_date": None,
            "author": None,
            "title": None,
            "sheet_count": 0,
            "sheet_names": [],
        }
        
        try:
            # Use openpyxl to get document properties
            wb = load_workbook(file_path, read_only=True)
            
            # Get sheet information
            metadata["sheet_count"] = len(wb.sheetnames)
            metadata["sheet_names"] = wb.sheetnames
            
            # Get document properties
            doc_props = self._extract_workbook_properties(wb)
            if doc_props:
                metadata.update({
                    "author": doc_props.get("creator"),
                    "title": doc_props.get("title"),
                    "subject": doc_props.get("subject"),
                    "creation_date": doc_props.get("created"),
                    "modification_date": doc_props.get("modified"),
                })
            
            # Close workbook to free resources
            wb.close()
            
        except Exception as e:
            logger.error(f"Error extracting metadata from Excel {file_path}: {str(e)}")
            raise
        
        return metadata

    def _extract_workbook_properties(self, workbook) -> Dict[str, Any]:
        """
        Extract document properties from an Excel workbook.
        
        Args:
            workbook: openpyxl Workbook object
            
        Returns:
            Dictionary containing document properties
        """
        properties = {}
        
        # Extract core properties if available
        if hasattr(workbook, 'properties') and workbook.properties:
            props = workbook.properties
            
            # Extract basic properties
            if hasattr(props, 'creator'):
                properties['creator'] = props.creator
            
            if hasattr(props, 'title'):
                properties['title'] = props.title
                
            if hasattr(props, 'subject'):
                properties['subject'] = props.subject
                
            if hasattr(props, 'description'):
                properties['description'] = props.description
                
            if hasattr(props, 'keywords'):
                properties['keywords'] = props.keywords
                
            if hasattr(props, 'category'):
                properties['category'] = props.category
            
            # Extract date information
            if hasattr(props, 'created') and props.created:
                properties['created'] = props.created.isoformat() if hasattr(props.created, 'isoformat') else str(props.created)
                
            if hasattr(props, 'modified') and props.modified:
                properties['modified'] = props.modified.isoformat() if hasattr(props.modified, 'isoformat') else str(props.modified)
        
        return properties
    
    def _extract_sheet_data(self, file_path: Union[str, Path], sheet_name: str) -> Dict[str, Any]:
        """
        Extract data from a single Excel sheet.
        
        Args:
            file_path: Path to the Excel document
            sheet_name: Name of the sheet to extract
            
        Returns:
            Dictionary containing sheet data
        """
        # Use pandas to read the Excel sheet
        df = pd.read_excel(file_path, sheet_name=sheet_name)
        
        # Replace NaN values with empty strings for easier processing
        df = df.replace({np.nan: ''})
        
        # Extract tables from the sheet
        tables = self._identify_tables_in_dataframe(df)
        
        # Construct sheet data dictionary
        sheet_data = {
            "name": sheet_name,
            "row_count": len(df),
            "column_count": len(df.columns),
            "tables": tables,
        }
        
        return sheet_data
    
    def _identify_tables_in_dataframe(self, df: pd.DataFrame) -> List[Dict[str, Any]]:
        """
        Identify tables within a pandas DataFrame.
        
        Args:
            df: pandas DataFrame representing the sheet
            
        Returns:
            List of dictionaries, each representing a table
        """
        tables = []
        
        # Skip empty DataFrames
        if df.empty:
            return tables
            
        # For now, treat the entire sheet as one table
        # In a more advanced implementation, we could detect multiple tables
        # based on empty rows/columns
        
        # Convert DataFrame to list of lists
        data = df.values.tolist()
        
        # Include column headers as first row
        headers = df.columns.tolist()
        data.insert(0, headers)
        
        # Determine if first row is a header
        has_header = self._detect_header_row(data)
        
        # Extract column names if header row exists
        columns = []
        if has_header and len(data) > 0:
            columns = data[0]
        
        tables.append({
            "data": data,
            "columns": columns,
            "has_header": has_header,
            "row_count": len(data),
            "column_count": len(data[0]) if data else 0,
        })
        
        return tables
    
    def _detect_header_row(self, table_data: List[List[Any]]) -> bool:
        """
        Detect if the first row of a table is a header row.
        
        Args:
            table_data: Table data as a list of rows
            
        Returns:
            True if the first row appears to be a header, False otherwise
        """
        if not table_data or len(table_data) < 2:
            return False
        
        # Check if the first row has different formatting or content pattern
        first_row = table_data[0]
        
        # Convert all elements to strings for consistent processing
        first_row = [str(cell) if cell is not None else "" for cell in first_row]
        
        # If all cells in the first row are capitalized or contain keywords
        header_indicators = ["total", "subtotal", "item", "description", "qty", "quantity", 
                            "amount", "price", "cost", "date", "no.", "number", "unit"]
        
        # Check for header-like formatting or content
        capitals_count = sum(1 for cell in first_row if cell and cell.isupper())
        indicators_count = sum(1 for cell in first_row if cell and 
                             any(indicator in cell.lower() for indicator in header_indicators))
        
        # Return True if first row has more capitalized cells or header indicators
        return (capitals_count > len(first_row) // 2) or (indicators_count > 0)
    
    def _is_financial_dataframe(self, df: pd.DataFrame) -> bool:
        """
        Determine if a DataFrame contains financial information.
        
        Args:
            df: pandas DataFrame
            
        Returns:
            True if the DataFrame appears to be financial, False otherwise
        """
        if df.empty:
            return False
            
        # Check for financial column headers
        financial_headers = ["amount", "total", "price", "cost", "subtotal", "balance", "sum", 
                            "payment", "claim", "value", "$", "usd", "dollar"]
        
        # Convert all column names to strings
        str_columns = [str(col).lower() for col in df.columns]
        
        # Check for financial terms in column names
        has_financial_header = any(
            any(fh in col for fh in financial_headers)
            for col in str_columns
        )
        
        if has_financial_header:
            return True
            
        # If no financial headers, check content for currency patterns
        currency_cells = 0
        total_cells = df.size
        
        # Convert DataFrame to string for easier regex processing
        str_df = df.astype(str)
        
        # Check each cell for currency patterns
        for col in str_df.columns:
            currency_cells += str_df[col].str.contains(r'(\$\s*[\d,.]+)|(\d{1,3}(?:,\d{3})*\.\d{2})').sum()
        
        # Calculate percentage of cells with currency values
        currency_cell_percentage = currency_cells / total_cells if total_cells > 0 else 0
        
        # Return True if significant currency cells are found
        return currency_cell_percentage > 0.15
    
    def _extract_financial_items_from_dataframe(
        self, df: pd.DataFrame, sheet_name: str, table_idx: int
    ) -> List[Dict[str, Any]]:
        """
        Extract financial line items from a DataFrame.
        
        Args:
            df: pandas DataFrame containing financial data
            sheet_name: Name of the sheet
            table_idx: Index of the table in the sheet
            
        Returns:
            List of dictionaries representing financial line items
        """
        line_items = []
        
        if df.empty:
            return line_items
            
        # Identify key columns
        description_col, amount_col, quantity_col, unit_price_col = self._identify_key_columns(df)
        
        # Convert column indices to names
        columns = df.columns.tolist()
        description_name = columns[description_col] if description_col is not None else None
        amount_name = columns[amount_col] if amount_col is not None else None
        quantity_name = columns[quantity_col] if quantity_col is not None else None
        unit_price_name = columns[unit_price_col] if unit_price_col is not None else None
        
        # Process each row
        for idx, row in df.iterrows():
            # Skip rows with NaN in key columns
            if (amount_col is not None and pd.isna(row[amount_name])) and \
               (description_col is not None and pd.isna(row[description_name])):
                continue
                
            # Skip rows that appear to be subtotals/totals
            if description_col is not None and isinstance(row[description_name], str):
                desc_lower = row[description_name].lower()
                if any(kw in desc_lower for kw in ["total", "subtotal", "sum"]):
                    continue
            
            # Create line item dictionary
            line_item = {
                "sheet_name": sheet_name,
                "table_index": table_idx,
                "row_number": idx + 1,  # 1-based row number
                "description": "",
                "amount": None,
                "quantity": None,
                "unit_price": None,
                "raw_data": row.to_dict(),
            }
            
            # Extract description
            if description_col is not None:
                desc_val = row[description_name]
                line_item["description"] = str(desc_val) if not pd.isna(desc_val) else ""
            
            # Extract amount
            if amount_col is not None:
                amount_val = row[amount_name]
                line_item["amount"] = self._extract_numeric_value(amount_val)
            
            # Extract quantity
            if quantity_col is not None:
                qty_val = row[quantity_name]
                line_item["quantity"] = self._extract_numeric_value(qty_val)
            
            # Extract unit price
            if unit_price_col is not None:
                unit_val = row[unit_price_name]
                line_item["unit_price"] = self._extract_numeric_value(unit_val)
            
            # Add line item if it has meaningful data
            if line_item["description"] or line_item["amount"] is not None:
                line_items.append(line_item)
        
        return line_items
    
    def _identify_key_columns(self, df: pd.DataFrame) -> Tuple[Optional[int], Optional[int], Optional[int], Optional[int]]:
        """
        Identify key columns in a financial table.
        
        Args:
            df: pandas DataFrame
            
        Returns:
            Tuple of (description_col, amount_col, quantity_col, unit_price_col) indices
        """
        description_col = None
        amount_col = None
        quantity_col = None
        unit_price_col = None
        
        # Get column names as strings
        columns = [str(col).lower() for col in df.columns]
        
        # Find description column
        description_keywords = ["description", "item", "work", "scope", "detail"]
        for idx, col in enumerate(columns):
            if any(kw in col for kw in description_keywords):
                description_col = idx
                break
                
        # If no description column found, try using the first text-heavy column
        if description_col is None:
            for idx, col in enumerate(df.columns):
                # Check if column contains mostly text values
                text_ratio = df[col].apply(lambda x: isinstance(x, str) and len(x) > 5).mean()
                if text_ratio > 0.5:  # If more than 50% of values are text
                    description_col = idx
                    break
        
        # Find amount column
        amount_keywords = ["amount", "total", "cost", "price", "value", "sum"]
        currency_counts = [0] * len(columns)
        
        # Count currency patterns in each column
        for idx, col in enumerate(df.columns):
            # Convert to string for pattern matching
            str_col = df[col].astype(str)
            # Count currency patterns
            currency_counts[idx] = str_col.str.contains(r'(\$\s*[\d,.]+)|(\d{1,3}(?:,\d{3})*\.\d{2})').sum()
        
        # First check keywords in column names
        for idx, col in enumerate(columns):
            if any(kw in col for kw in amount_keywords):
                amount_col = idx
                break
                
        # If no amount column found by name, use the column with most currency patterns
        if amount_col is None and currency_counts:
            max_count = max(currency_counts)
            if max_count > 0:
                amount_col = currency_counts.index(max_count)
        
        # Find quantity column
        quantity_keywords = ["qty", "quantity", "count", "units"]
        for idx, col in enumerate(columns):
            if any(kw in col for kw in quantity_keywords):
                quantity_col = idx
                break
        
        # Find unit price column
        unit_price_keywords = ["unit price", "unit cost", "rate", "price per"]
        for idx, col in enumerate(columns):
            if any(kw in col for kw in unit_price_keywords):
                unit_price_col = idx
                break
        
        return description_col, amount_col, quantity_col, unit_price_col
    
    def _extract_numeric_value(self, value: Any) -> Optional[float]:
        """
        Extract numeric value from a cell value.
        
        Args:
            value: Cell value from Excel
            
        Returns:
            Extracted float value or None if no valid value found
        """
        if pd.isna(value):
            return None
            
        # If already a number, return it
        if isinstance(value, (int, float)):
            return float(value)
            
        # If datetime, return None
        if isinstance(value, (datetime, pd.Timestamp)):
            return None
            
        # Convert to string and remove currency symbols and commas
        if not isinstance(value, str):
            value = str(value)
            
        # Remove currency symbols and commas
        cleaned_text = value.replace('$', '').replace(',', '').strip()
        
        # Extract numeric value
        try:
            # Find the first sequence of digits with optional decimal point
            matches = re.search(r'-?\d+\.?\d*', cleaned_text)
            if matches:
                return float(matches.group(0))
            return None
        except (ValueError, TypeError):
            return None