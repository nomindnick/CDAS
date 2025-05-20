"""
Excel exporter for the Construction Document Analysis System.

This module provides functionality for exporting financial data to Excel formats.
"""

import logging
import tempfile
from typing import Dict, List, Any, Optional, Union
from datetime import datetime
from pathlib import Path
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from cdas.db.models import LineItem, Document
from cdas.financial_analysis.engine import FinancialAnalysisEngine

# Set up logging
logger = logging.getLogger(__name__)


class ExcelExporter:
    """Exports financial data to Excel format."""
    
    def __init__(self, db_session: Session):
        """Initialize the Excel exporter.
        
        Args:
            db_session: Database session
        """
        self.db_session = db_session
        self.analysis_engine = FinancialAnalysisEngine(db_session)
        
    def export_financial_summary(self, output_path: str, project_id: Optional[str] = None) -> Dict[str, Any]:
        """Export a financial summary to Excel.
        
        Args:
            output_path: Path to save the Excel file
            project_id: Optional project ID to filter data
            
        Returns:
            Dictionary with export information
        """
        logger.info(f"Exporting financial summary to Excel: {output_path}")
        
        # Create a writer to save the Excel file
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # Export various data sheets
            self._export_summary_sheet(writer, project_id)
            self._export_line_items_sheet(writer, project_id)
            self._export_disputed_amounts_sheet(writer, project_id)
            self._export_documents_sheet(writer, project_id)
            
            # Get the workbook to apply formatting
            workbook = writer.book
            
            # Apply formatting to all sheets
            for sheet_name in workbook.sheetnames:
                self._format_sheet(workbook[sheet_name])
                
            # Auto-adjust column widths
            for sheet_name in workbook.sheetnames:
                self._auto_adjust_columns(workbook[sheet_name])
        
        return {
            'path': output_path,
            'sheets': ['Summary', 'Line Items', 'Disputed Amounts', 'Documents'],
            'generated_at': datetime.now().isoformat()
        }
        
    def export_analysis_results(self, output_path: str, project_id: Optional[str] = None) -> Dict[str, Any]:
        """Export analysis results to Excel.
        
        Args:
            output_path: Path to save the Excel file
            project_id: Optional project ID to filter data
            
        Returns:
            Dictionary with export information
        """
        logger.info(f"Exporting analysis results to Excel: {output_path}")
        
        # Create a writer to save the Excel file
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # Export analysis data sheets
            self._export_patterns_sheet(writer, project_id)
            self._export_anomalies_sheet(writer, project_id)
            self._export_relationships_sheet(writer, project_id)
            
            # Get the workbook to apply formatting
            workbook = writer.book
            
            # Apply formatting to all sheets
            for sheet_name in workbook.sheetnames:
                self._format_sheet(workbook[sheet_name])
                
            # Auto-adjust column widths
            for sheet_name in workbook.sheetnames:
                self._auto_adjust_columns(workbook[sheet_name])
        
        return {
            'path': output_path,
            'sheets': ['Patterns', 'Anomalies', 'Relationships'],
            'generated_at': datetime.now().isoformat()
        }
        
    def export_amount_analysis(self, amount: float, output_path: str) -> Dict[str, Any]:
        """Export analysis of a specific amount to Excel.
        
        Args:
            amount: Amount to analyze
            output_path: Path to save the Excel file
            
        Returns:
            Dictionary with export information
        """
        logger.info(f"Exporting amount analysis to Excel: {output_path}")
        
        # Analyze the amount
        analysis_results = self.analysis_engine.analyze_amount(amount)
        matches = analysis_results.get('matches', [])
        anomalies = analysis_results.get('anomalies', [])
        
        # Create a writer to save the Excel file
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # Create a summary sheet
            summary_df = pd.DataFrame({
                'Item': ['Amount', 'Number of Matches', 'Number of Anomalies'],
                'Value': [f"${amount:,.2f}", len(matches), len(anomalies)]
            })
            summary_df.to_excel(writer, sheet_name='Summary', index=False)
            
            # Create a matches sheet
            if matches:
                matches_df = pd.DataFrame(matches)
                matches_df.to_excel(writer, sheet_name='Matches', index=False)
            
            # Create an anomalies sheet
            if anomalies:
                anomalies_df = pd.DataFrame(anomalies)
                anomalies_df.to_excel(writer, sheet_name='Anomalies', index=False)
                
            # Get the workbook to apply formatting
            workbook = writer.book
            
            # Apply formatting to all sheets
            for sheet_name in workbook.sheetnames:
                self._format_sheet(workbook[sheet_name])
                
            # Auto-adjust column widths
            for sheet_name in workbook.sheetnames:
                self._auto_adjust_columns(workbook[sheet_name])
        
        return {
            'path': output_path,
            'amount': amount,
            'match_count': len(matches),
            'anomaly_count': len(anomalies),
            'generated_at': datetime.now().isoformat()
        }
    
    def _export_summary_sheet(self, writer: pd.ExcelWriter, project_id: Optional[str] = None) -> None:
        """Export the summary sheet.
        
        Args:
            writer: Excel writer
            project_id: Optional project ID to filter data
        """
        # Get summary data
        from cdas.db.operations import get_documents, get_line_items
        
        # Get documents and line items
        documents = get_documents(self.db_session, project_id=project_id)
        line_items = get_line_items(self.db_session, project_id=project_id)
        
        # Calculate summary statistics
        total_documents = len(documents)
        total_items = len(line_items)
        
        # Calculate total amount
        total_amount = sum(
            float(item.amount) 
            for item in line_items 
            if item.amount is not None
        )
        
        # Calculate disputed amount
        disputed_amounts = self.analysis_engine.find_disputed_amounts()
        total_disputed = sum(
            amount.get('amount', 0) 
            for amount in disputed_amounts 
            if amount.get('amount') is not None
        )
        
        # Get document counts by type
        doc_by_type = {}
        for doc in documents:
            doc_type = doc.doc_type or 'unknown'
            if doc_type not in doc_by_type:
                doc_by_type[doc_type] = 0
            doc_by_type[doc_type] += 1
        
        # Get document counts by party
        doc_by_party = {}
        for doc in documents:
            party = doc.party or 'unknown'
            if party not in doc_by_party:
                doc_by_party[party] = 0
            doc_by_party[party] += 1
        
        # Create summary DataFrame
        summary_data = {
            'Metric': [
                'Total Documents', 
                'Total Line Items', 
                'Total Amount', 
                'Total Disputed Amount',
                'Disputed Percentage'
            ],
            'Value': [
                total_documents,
                total_items,
                f"${total_amount:,.2f}",
                f"${total_disputed:,.2f}",
                f"{(total_disputed / total_amount * 100) if total_amount else 0:.2f}%"
            ]
        }
        
        summary_df = pd.DataFrame(summary_data)
        
        # Create document type and party DataFrames
        doc_type_df = pd.DataFrame({
            'Document Type': list(doc_by_type.keys()),
            'Count': list(doc_by_type.values())
        })
        
        doc_party_df = pd.DataFrame({
            'Party': list(doc_by_party.keys()),
            'Count': list(doc_by_party.values())
        })
        
        # Write to Excel
        summary_df.to_excel(writer, sheet_name='Summary', startrow=1, startcol=0, index=False)
        
        # Add document type stats
        doc_type_df.to_excel(writer, sheet_name='Summary', startrow=len(summary_df) + 4, startcol=0, index=False)
        
        # Add document party stats
        doc_party_df.to_excel(writer, sheet_name='Summary', startrow=len(summary_df) + len(doc_type_df) + 7, startcol=0, index=False)
        
        # Get the worksheet to add a title
        worksheet = writer.sheets['Summary']
        worksheet.cell(row=1, column=1, value="Financial Summary Report")
        worksheet.cell(row=1, column=1).font = Font(size=14, bold=True)
        
    def _export_line_items_sheet(self, writer: pd.ExcelWriter, project_id: Optional[str] = None) -> None:
        """Export line items to a sheet.
        
        Args:
            writer: Excel writer
            project_id: Optional project ID to filter data
        """
        from cdas.db.operations import get_line_items
        
        # Get line items
        line_items = get_line_items(self.db_session, project_id=project_id)
        
        # Convert to list of dictionaries for DataFrame
        items_data = []
        for item in line_items:
            # Format amount as string with currency symbol
            amount_str = f"${float(item.amount):,.2f}" if item.amount is not None else "N/A"
            
            items_data.append({
                'Item ID': item.item_id,
                'Document ID': item.doc_id,
                'Description': item.description,
                'Amount': amount_str,
                'Quantity': item.quantity,
                'Unit': item.unit,
                'Date': item.date.strftime('%Y-%m-%d') if item.date else 'N/A',
                'Page Number': item.page_num
            })
        
        # Create DataFrame
        if items_data:
            items_df = pd.DataFrame(items_data)
            items_df.to_excel(writer, sheet_name='Line Items', index=False)
        else:
            # Create an empty DataFrame with column headers
            items_df = pd.DataFrame(columns=[
                'Item ID', 'Document ID', 'Description', 'Amount', 
                'Quantity', 'Unit', 'Date', 'Page Number'
            ])
            items_df.to_excel(writer, sheet_name='Line Items', index=False)
            
    def _export_disputed_amounts_sheet(self, writer: pd.ExcelWriter, project_id: Optional[str] = None) -> None:
        """Export disputed amounts to a sheet.
        
        Args:
            writer: Excel writer
            project_id: Optional project ID to filter data
        """
        # Get disputed amounts
        disputed_amounts = self.analysis_engine.find_disputed_amounts()
        
        # Convert to DataFrame
        if disputed_amounts:
            # Format each amount entry
            for amount in disputed_amounts:
                if 'amount' in amount:
                    amount['amount'] = f"${float(amount['amount']):,.2f}"
            
            disputed_df = pd.DataFrame(disputed_amounts)
            disputed_df.to_excel(writer, sheet_name='Disputed Amounts', index=False)
        else:
            # Create an empty DataFrame with column headers
            disputed_df = pd.DataFrame(columns=[
                'amount', 'description', 'doc_id', 'party', 'explanation'
            ])
            disputed_df.to_excel(writer, sheet_name='Disputed Amounts', index=False)
    
    def _export_documents_sheet(self, writer: pd.ExcelWriter, project_id: Optional[str] = None) -> None:
        """Export documents to a sheet.
        
        Args:
            writer: Excel writer
            project_id: Optional project ID to filter data
        """
        from cdas.db.operations import get_documents
        
        # Get documents
        documents = get_documents(self.db_session, project_id=project_id)
        
        # Convert to list of dictionaries for DataFrame
        docs_data = []
        for doc in documents:
            docs_data.append({
                'Document ID': doc.doc_id,
                'File Name': doc.file_name,
                'Document Type': doc.doc_type,
                'Party': doc.party,
                'Date': doc.date.strftime('%Y-%m-%d') if doc.date else 'N/A',
                'Page Count': doc.page_count,
                'Status': doc.status
            })
        
        # Create DataFrame
        if docs_data:
            docs_df = pd.DataFrame(docs_data)
            docs_df.to_excel(writer, sheet_name='Documents', index=False)
        else:
            # Create an empty DataFrame with column headers
            docs_df = pd.DataFrame(columns=[
                'Document ID', 'File Name', 'Document Type', 'Party', 
                'Date', 'Page Count', 'Status'
            ])
            docs_df.to_excel(writer, sheet_name='Documents', index=False)
    
    def _export_patterns_sheet(self, writer: pd.ExcelWriter, project_id: Optional[str] = None) -> None:
        """Export patterns to a sheet.
        
        Args:
            writer: Excel writer
            project_id: Optional project ID to filter data
        """
        # Get patterns
        suspicious_patterns = self.analysis_engine.find_suspicious_patterns()
        
        # Convert to DataFrame
        if suspicious_patterns:
            # Process pattern data
            pattern_data = []
            for pattern in suspicious_patterns:
                # Convert doc_ids list to string if present
                if 'doc_ids' in pattern:
                    pattern['documents'] = ', '.join(pattern['doc_ids'])
                
                # Format confidence as percentage
                if 'confidence' in pattern:
                    pattern['confidence'] = f"{pattern['confidence'] * 100:.2f}%"
                    
                pattern_data.append(pattern)
                
            patterns_df = pd.DataFrame(pattern_data)
            patterns_df.to_excel(writer, sheet_name='Patterns', index=False)
        else:
            # Create an empty DataFrame with column headers
            patterns_df = pd.DataFrame(columns=[
                'pattern_type', 'explanation', 'confidence', 'documents'
            ])
            patterns_df.to_excel(writer, sheet_name='Patterns', index=False)
    
    def _export_anomalies_sheet(self, writer: pd.ExcelWriter, project_id: Optional[str] = None) -> None:
        """Export anomalies to a sheet.
        
        Args:
            writer: Excel writer
            project_id: Optional project ID to filter data
        """
        # Get anomalies
        anomalies = self.analysis_engine.find_anomalies()
        
        # Convert to DataFrame
        if anomalies:
            # Process anomaly data
            anomaly_data = []
            for anomaly in anomalies:
                # Format confidence as percentage
                if 'confidence' in anomaly:
                    anomaly['confidence'] = f"{anomaly['confidence'] * 100:.2f}%"
                    
                anomaly_data.append(anomaly)
                
            anomalies_df = pd.DataFrame(anomaly_data)
            anomalies_df.to_excel(writer, sheet_name='Anomalies', index=False)
        else:
            # Create an empty DataFrame with column headers
            anomalies_df = pd.DataFrame(columns=[
                'anomaly_type', 'explanation', 'confidence'
            ])
            anomalies_df.to_excel(writer, sheet_name='Anomalies', index=False)
    
    def _export_relationships_sheet(self, writer: pd.ExcelWriter, project_id: Optional[str] = None) -> None:
        """Export document relationships to a sheet.
        
        Args:
            writer: Excel writer
            project_id: Optional project ID to filter data
        """
        from cdas.db.models import DocumentRelationship
        
        # Query all relationships
        relationships_query = self.db_session.query(DocumentRelationship)
        
        # If project_id is provided, filter by documents in the project
        if project_id:
            from cdas.db.operations import get_documents
            documents = get_documents(self.db_session, project_id=project_id)
            doc_ids = [doc.doc_id for doc in documents]
            
            relationships_query = relationships_query.filter(
                DocumentRelationship.source_doc_id.in_(doc_ids) |
                DocumentRelationship.target_doc_id.in_(doc_ids)
            )
        
        # Get all relationships
        relationships = relationships_query.all()
        
        # Convert to list of dictionaries for DataFrame
        rel_data = []
        for rel in relationships:
            rel_data.append({
                'Source Document': rel.source_doc_id,
                'Target Document': rel.target_doc_id,
                'Relationship Type': rel.relationship_type,
                'Confidence': f"{float(rel.confidence) * 100:.2f}%" if rel.confidence else "100%"
            })
        
        # Create DataFrame
        if rel_data:
            rel_df = pd.DataFrame(rel_data)
            rel_df.to_excel(writer, sheet_name='Relationships', index=False)
        else:
            # Create an empty DataFrame with column headers
            rel_df = pd.DataFrame(columns=[
                'Source Document', 'Target Document', 'Relationship Type', 'Confidence'
            ])
            rel_df.to_excel(writer, sheet_name='Relationships', index=False)
    
    def _format_sheet(self, worksheet: openpyxl.worksheet.worksheet.Worksheet) -> None:
        """Apply formatting to a worksheet.
        
        Args:
            worksheet: The worksheet to format
        """
        # Style for headers (first row)
        header_font = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Style for data cells
        data_font = Font(size=11)
        data_alignment = Alignment(vertical='center', wrap_text=True)
        
        # Borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Apply styles to header row
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Apply styles to data rows
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.font = data_font
                cell.alignment = data_alignment
                cell.border = thin_border
                
        # Apply alternating row colors
        for row_idx, row in enumerate(worksheet.iter_rows(min_row=2)):
            if row_idx % 2:
                for cell in row:
                    cell.fill = PatternFill(start_color="EDF3F7", end_color="EDF3F7", fill_type="solid")
    
    def _auto_adjust_columns(self, worksheet: openpyxl.worksheet.worksheet.Worksheet) -> None:
        """Auto-adjust column widths based on content.
        
        Args:
            worksheet: The worksheet to adjust
        """
        dims = {}
        for row in worksheet.rows:
            for cell in row:
                if cell.value:
                    dims[cell.column_letter] = max(
                        (dims.get(cell.column_letter, 0)),
                        len(str(cell.value)) + 2
                    )
        
        # Set column widths
        for col, width in dims.items():
            # Limit max width to 50 characters
            worksheet.column_dimensions[col].width = min(width, 50)